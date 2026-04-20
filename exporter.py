import io, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLON_HDR_BG = '1F4E79'
WH_HDR_BG    = '0F6E56'
DATA_HDR_BG  = 'D6E4F0'
DATA_HDR_BG_WH='D1F0E8'
TOTAL_BG     = 'EFF3F8'
ALT_BG       = 'F7FAFD'
WHITE        = 'FFFFFF'

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value=None, bold=False, size=9, color='000000',
         bg=None, align='left', valign='center', wrap=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Calibri', size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    c.border = thin_border()
    if bg: c.fill = PatternFill('solid', start_color=bg)
    if num_fmt: c.number_format = num_fmt
    return c

def set_col_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

COLON_COLS  = ['Código','Descripción','Cantidad','Marca','L (cm)','W (cm)','H (cm)','PZxB','Peso (kg)','Vol. (m³)','Bultos','N° Bultos','Ubicación']
COLON_W     = [18,46,10,12,8,8,8,7,12,12,9,12,12]
WH_COLS     = ['Código','Descripción','Cantidad','L (cm)','W (cm)','H (cm)','PZxB','kg','m3','Bultos','N° Bultos','Desde','Cód. barras']
WH_W        = [20,48,10,8,8,8,7,10,10,8,12,18,14]

def write_entrega_block(ws, row, e):
    is_wh = e['tipo'] == 'WH'
    bg    = WH_HDR_BG if is_wh else COLON_HDR_BG
    dHdrBg= DATA_HDR_BG_WH if is_wh else DATA_HDR_BG
    cols  = WH_COLS if is_wh else COLON_COLS
    nc    = len(cols)

    # Header row 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1,
                value=f"{'Albarán' if is_wh else 'Entrega'}: {e['entrega']}   |   "
                      f"Orden: {e['orden']}   |   "
                      f"{'Cita agendada' if is_wh else 'Fecha'}: {e['fecha']}   |   Estado: Hecho")
    c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin_border()
    ws.row_dimensions[row].height = 20
    row += 1

    # Header row 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1,
                value=f"Cliente: {e.get('cliente','')}   |   Vendedor: {e.get('vendedor','')}   |   Almacén: {e.get('almacen','')}")
    c.font = Font(name='Calibri', size=9, color='D0E8FF')
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin_border()
    ws.row_dimensions[row].height = 14
    row += 1

    # Column headers
    for ci, h in enumerate(cols, 1):
        cell(ws, row, ci, h, bold=True, size=8, bg=dHdrBg, align='center', wrap=True)
    ws.row_dimensions[row].height = 22
    row += 1

    # Data
    tkg = tm3 = tqty = tbultos = 0
    for ri, it in enumerate(e['items']):
        bg_row = ALT_BG if ri % 2 else WHITE
        dim_ok = any(it.get(k) for k in ['l','w','h'])
        if is_wh:
            vals = [it.get('codigo',''), it.get('descripcion',''), it.get('cantidad',''),
                    it.get('l','') if dim_ok else '', it.get('w','') if dim_ok else '',
                    it.get('h','') if dim_ok else '', it.get('pzb',''),
                    it.get('kg',''), it.get('m3',''), it.get('bultos','') or '',
                    it.get('nbultos',''), it.get('desde',''), it.get('barcode','')]
        else:
            vals = [it.get('codigo',''), it.get('descripcion',''), it.get('cantidad',''),
                    it.get('marca',''),
                    it.get('l','') if dim_ok else '', it.get('w','') if dim_ok else '',
                    it.get('h','') if dim_ok else '', it.get('pzb',''),
                    it.get('kg',''), it.get('m3',''), it.get('bultos','') or '',
                    it.get('nbultos',''), it.get('ubicacion','')]
        for ci, v in enumerate(vals, 1):
            aln = 'right' if ci in (3,5,6,7,8,9,10,11) else ('center' if ci == 4 else 'left')
            cell(ws, row, ci, v, size=8, bg=bg_row, align=aln, wrap=(ci==2))
        ws.row_dimensions[row].height = 14
        tkg += it.get('kg') or 0; tm3 += it.get('m3') or 0
        tqty += it.get('cantidad') or 0; tbultos += it.get('bultos') or 0
        row += 1

    # Total
    if is_wh:
        tots = ['TOTAL','',tqty,'','','','',round(tkg,3) or '',round(tm3,4) or '',tbultos or '','','','']
    else:
        tots = ['TOTAL','',tqty,'','','','','',round(tkg,3) or '',round(tm3,4) or '',tbultos or '','','']
    for ci, v in enumerate(tots, 1):
        cell(ws, row, ci, v, bold=True, size=8, bg=TOTAL_BG, align='right' if ci>2 else 'left')
    ws.row_dimensions[row].height = 14
    return row + 2

def export_to_excel(entregas, out):
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Vista PDF
    ws1 = wb.create_sheet('Vista PDF')
    ws1.sheet_view.showGridLines = False
    set_col_width(ws1, COLON_W)
    row = 1
    for e in entregas:
        row = write_entrega_block(ws1, row, e)
    ws1.freeze_panes = 'A1'

    # Sheet 2: Resumen
    ws2 = wb.create_sheet('Resumen')
    ws2.sheet_view.showGridLines = False
    all_items = [it for e in entregas for it in e['items']]
    tkg = sum(it.get('kg') or 0 for it in all_items)
    tm3 = sum(it.get('m3') or 0 for it in all_items)
    tb  = sum(e.get('bultos') or 0 for e in entregas)

    ws2.merge_cells('A1:K1')
    c = ws2.cell(row=1, column=1, value='RESUMEN CONSOLIDADO')
    c.font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    c.fill = PatternFill('solid', start_color=COLON_HDR_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
    ws2.row_dimensions[1].height = 24

    ent_hdrs = ['Tipo','Entrega','Orden','Fecha','Cliente','Líneas','Unidades','Peso (kg)','Vol. (m³)','Bultos','Marcas']
    for ci, h in enumerate(ent_hdrs, 1):
        cell(ws2, 2, ci, h, bold=True, size=8, bg=DATA_HDR_BG, align='center')
    ws2.row_dimensions[2].height = 16

    for ri, e in enumerate(entregas):
        bg = ALT_BG if ri % 2 else WHITE
        items = e['items']
        ekg = sum(it.get('kg') or 0 for it in items)
        em3 = sum(it.get('m3') or 0 for it in items)
        eq  = sum(it.get('cantidad') or 0 for it in items)
        marcas = ', '.join(sorted(set(it.get('marca','') for it in items if it.get('marca'))))
        vals = [e['tipo'], e['entrega'], e['orden'], e['fecha'], e.get('cliente',''),
                len(items), eq, round(ekg,3), round(em3,4), e.get('bultos',0), marcas]
        for ci, v in enumerate(vals, 1):
            cell(ws2, ri+3, ci, v, size=8, bg=bg, align='right' if ci in (6,7,8,9,10) else 'left')
        ws2.row_dimensions[ri+3].height = 14

    tr = len(entregas)+3
    tots2 = ['','TOTAL','','','',len(all_items),
             sum(it.get('cantidad') or 0 for it in all_items),
             round(tkg,3), round(tm3,4), tb, '']
    for ci, v in enumerate(tots2, 1):
        cell(ws2, tr, ci, v, bold=True, size=8, bg=TOTAL_BG, align='right' if ci in (6,7,8,9,10) else 'left')
    ws2.row_dimensions[tr].height = 14
    set_col_width(ws2, [7,22,10,12,24,8,10,12,12,8,20])
    ws2.freeze_panes = 'A2'

    # Sheet 3: Base de datos
    ws3 = wb.create_sheet('Base de datos')
    ws3.sheet_view.showGridLines = False
    hdrs3 = ['Tipo','Entrega','Orden','Fecha','Archivo','Cliente','Código','Descripción',
              'Cantidad','Marca','L(cm)','W(cm)','H(cm)','PZxB','kg','m³','Bultos','N°Bultos','Ubicación/Desde']
    for ci, h in enumerate(hdrs3, 1):
        cell(ws3, 1, ci, h, bold=True, size=8, bg=COLON_HDR_BG, color=WHITE, align='center')
    ws3.row_dimensions[1].height = 20

    for ri, (e, it) in enumerate([(e, it) for e in entregas for it in e['items']]):
        bg = ALT_BG if ri % 2 else WHITE
        vals = [e['tipo'], e['entrega'], e['orden'], e['fecha'], e.get('fileName',''),
                e.get('cliente',''), it.get('codigo',''), it.get('descripcion',''),
                it.get('cantidad',''), it.get('marca',''),
                it.get('l',''), it.get('w',''), it.get('h',''), it.get('pzb',''),
                it.get('kg',''), it.get('m3',''), it.get('bultos','') or '', it.get('nbultos',''),
                it.get('ubicacion','') or it.get('desde','')]
        for ci, v in enumerate(vals, 1):
            cell(ws3, ri+2, ci, v, size=8, bg=bg,
                 align='right' if ci in (9,11,12,13,14,15,16,17) else 'left')
        ws3.row_dimensions[ri+2].height = 13

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(hdrs3))}1"
    set_col_width(ws3, [7,22,10,12,20,22,18,44,10,12,8,8,8,7,10,10,8,12,18])
    ws3.freeze_panes = 'A2'

    if isinstance(out, str):
        wb.save(out)
    else:
        wb.save(out)
